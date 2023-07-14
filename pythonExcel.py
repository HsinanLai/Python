from dataclasses import is_dataclass
from pickle import TRUE
from winreg import ExpandEnvironmentStrings
import paramiko
import os
import sys
import shutil
from os import listdir
from os.path import isfile, isdir, join
from datetime import datetime
from pathlib import Path
from configparser import ConfigParser
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from zipfile import ZipFile
import smtplib
import time
import fnmatch
import csv
import sqlite3
import openpyxl
#import datetime

#function : log2file
def log2file(strlog):
    Conf = ConfigParser()	
    Conf.read("pyxl2ai.ini",encoding="utf-8")
    Log = Conf['Log']
    SFTP = Conf['SFTP']
    FileInfo = Conf['FileInfo']
    PLine = FileInfo['PLine']
    now = datetime.now() # current date and time
    datestr = now.strftime("%Y%m%d")
    date_time = now.strftime("%Y/%m/%d, %H:%M:%S.%f")
    logpath = Log['LogPath']
    logfilename = 'BubbleAnalysis-' + PLine + '-' + datestr + '.log' 
    logfile = join( logpath , logfilename )
    print(strlog)
    lf = open(logfile,'a+',encoding='UTF8', newline='')
    lf.write( date_time + ">>" + strlog +'\n' )
    lf.close()

#function : sendmail 
def sendmail(mailmsg):
    now = datetime.now() # current date and time
    date_time = now.strftime("%Y/%m/%d, %H:%M:%S.%f")
    # Step 2 - Create message object instance
    msg = MIMEMultipart()
    # Step 3 - Create message body
    message =  date_time + ">>>" + mailmsg
    MConf = ConfigParser()	
    MConf.read("pyxl2ai.ini",encoding="utf-8")
    FileInfo = MConf['FileInfo']
    PLine = FileInfo['PLine']
    SMTP = MConf['SMTP']
    # Step 4 - Declare SMTP credentials
    password = SMTP['SMTPpass'] 
    username = SMTP['SMTPuser'] #""
    smtphost = SMTP['SMTPhost'] #"smtp.office365.com:587"

    # 控制是否需要發送郵件
    EmailNotify = SMTP['EmailNotify']  #""
    if ( EmailNotify != 'Yes' ): 
        print('no need notify')
        return True

    # Step 5 - Declare message elements
    msg['From'] = SMTP['SMTPFrom']  #""
    recstr = SMTP['SMTPTo']    #""
    recipent= recstr.split(',')
    msg['To'] = ", ".join(recipent)
    Log = MConf['Log']
    now = datetime.now() # current date and time
    datestr = now.strftime("%Y%m%d")
    logpath = Log['LogPath']
    logfilename = 'BubbleAnalysis-' + PLine + '-' + datestr + '.log' 
    logfile = join( logpath , logfilename )
    #msg.attach(MIMEText(open(logfile).read()))
    
    # 構造附件
    att = MIMEText(open(logfile, "rb").read(), "base64", "utf-8")
    att["Content-Type"] = "application/octet-stream"
    # 附件名稱為中文時的寫法
    att.add_header("Content-Disposition", "attachment", filename=("utf-8", "", logfile))
    # 附件名稱非中文時的寫法
    # att["Content-Disposition"] = ‘attachment; filename="test.html")‘
    msg.attach(att)
    
    msg['Subject'] = "[IT Batch Program] " + PLine + " - " + date_time + ": Bubble Analysis for DataBaseServerComplete "
    # Step 6 - Add the message body to the object instance
    msg.attach(MIMEText(message, 'plain'))
    # Step 7 - Create the server connection
    server = smtplib.SMTP(smtphost)
    # Step 8 - Switch the connection over to TLS encryption
    server.starttls() 
    # Step 9 - Authenticate with the server
    server.login(username, password)
    # Step 10 - Send the message
    # server.sendmail(msg['From'], msg['To'], msg.as_string())
    server.sendmail(msg['From'], recipent, msg.as_string())
    # Step 11 - Disconnect
    server.quit()

#function : sftp 2 SFTP Server
def sftp(poutputfilename):
    now = datetime.now() # current date and time
    datestr = now.strftime("%Y%m%d")
    #Read Config data
    Config1 = ConfigParser()	
    Config1.read("pyxl2ai.ini",encoding="utf-8")
    SFTP = Config1['SFTP']
    SFTPip = SFTP['SFTPip']
    SFTPUser = SFTP['SFTPUser']
    SFTPPasswd = SFTP['SFTPPasswd']
    SFTPPort = SFTP['SFTPPort']
    FileInfo= Config1['FileInfo']
    OutputDir = FileInfo['OutputDir']
    OutputFileName = poutputfilename     #"//192.168.66.50/HotDataAI$"
    SFTPtargetpath = FileInfo['SFTPTargetPath']   #"/process_data/testftp/"
    #PLine = FileInfo['PLine']   #"/process_data/testftp/"
    outputfile = join( OutputDir , OutputFileName )
    exeStatus=True

    count = 1
    while True:
        try:
            client = paramiko.SSHClient()  
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            time.sleep(3) 
            client.connect(SFTPip,  username=SFTPUser, password=SFTPPasswd, port=SFTPPort, timeout=10 ) 
            transport = client.get_transport()  
            sftp = paramiko.SFTPClient.from_transport(transport)
            print("sftp connected")
            break
        except Exception as e:
            #print(e)
            log2file(' error:' + e.args[0] + " , retry:" + str(count) )
            if count < 5: 
                time.sleep(1) 
                print("retry>>>" , count)
                count = count+1 
                client.close()
                continue
            else:
                log2file(  ' error:' + e.args[0]  )
                client.close()
                exeStatus=False
                return False
                break
    if exeStatus: 
        log2file(  'SFTP connected successful' )
    try: 
        targertfile = join(SFTPtargetpath, OutputFileName)
        sftp.put(outputfile, targertfile)
        log2file( OutputFileName + ' put to SFTP server successful!!'  )
    except Exception as e:
        log2file( ' error:' + e.args[0]  )
        exeStatus=False
        client.close()
        return False
    client.close()
    return True


#main program
#讀取來源檔案的配置設定
Config = ConfigParser()	
Config.read("pyxl2ai.ini",encoding="utf-8")
FileInfo = Config['FileInfo']
SourceDir = FileInfo['SourceDir']
SourceFileName = FileInfo['SourceFileName']
SourceWorkSheet = FileInfo['SourceWorkSheet']
Tempdir = FileInfo['Tempdir']
TempFileName = FileInfo['TempFileName']
TZ = FileInfo['TZ']   #"/process_data/testftp/"
now = datetime.now() # current date and time
datetimestr = now.strftime("%Y%m%d%H%M%S") 
SourceFile = join(SourceDir,SourceFileName)
Tempfile = Tempdir + TempFileName+ datetimestr+ ".xlsx"
shutil.copyfile(SourceFile, Tempfile)
print(SourceFile)
#Get SourceFile data
FileSize = str(os.path.getsize(SourceFile))
filecrdt = datetime.fromtimestamp(os.path.getmtime(SourceFile))
filemtdt = datetime.fromtimestamp(os.path.getctime(SourceFile))
filecrdtstr = filecrdt.strftime("%m%d%Y %H%M%S")
filemtdtstr = filemtdt.strftime("%m%d%Y %H%M%S")
#Get last SourceFile Data
LastFileInfo = Config['LastFileInfo']
LastFileSize = LastFileInfo['filesize']
Lastfilecrdt = LastFileInfo['filecrdt']
Lastfilemtdt = LastFileInfo['filemtdt']
LastTestDate = LastFileInfo['TestDate']

if (FileSize == LastFileSize and filemtdtstr == Lastfilemtdt and filecrdtstr == Lastfilecrdt):
    log2file("same as the last file")
    sendmail('same as the last file')
    exit(0)
Config.set("LastFileInfo","FileName", SourceFile) 
Config.set("LastFileInfo","FileSize", FileSize) 
Config.set("LastFileInfo","filecrdt", filecrdtstr) 
Config.set("LastFileInfo","filemtdt", filemtdtstr) 


# 寫入 INI 檔案
with open('pyxl2ai.ini', 'w',encoding='UTF8') as configfile:
    Config.write(configfile)

c1 = sqlite3.connect("file::memory:?cache=shared", uri=True)
sqlstr = 'CREATE TABLE BubbleAnalysis(Sampling_Time,Location,Depth,Lon_Trail,short_diameter,Bubble_Thickness,Volume,Lon_Trail_Log,Internal_Pressure_Log,Internal_Pressure,N2,CO2,O2,Ar,SO2)'
c1.execute(sqlstr)
print("create database/table complete")
# 使用 load_workbook 讀取 test.xlsx
workbook = openpyxl.load_workbook(Tempfile, data_only=True)
exeStatus = TRUE
# 取得第一個工作表
sheet = workbook['BubbleDATA']

# 顯示 row總數 及 column總數
print('row總數:', sheet.max_row)
print('column總數:', sheet.max_column)

# 顯示 cell 資料
cnt=0
for i in range(1, sheet.max_row+1):
    #Data Read 
    AnaStatus = sheet.cell(row = i, column = 13).value #AnaStatus O:will display
    if (AnaStatus!="○"):
        continue
    TestDate = str(sheet.cell(row = i, column = 11).value) #Sampling_Time
    TestDatestr = TestDate[:10]
    #print(TestDatestr)
    TestDateDate = datetime. strptime(TestDatestr, '%Y-%m-%d')
    LastTestDateDate = datetime. strptime(LastTestDate, '%Y-%m-%d')
    if (TestDateDate <= LastTestDateDate):
        continue
    #print(TestDateDate)
   
    Sampling_Time = str(sheet.cell(row = i, column = 9).value) #Sampling_Time
    print(i,'sampletime:',Sampling_Time[:19])
    if (Sampling_Time == "" or Sampling_Time == 'None' or Sampling_Time == 'NA'):
        continue
    date_object = datetime.strptime(Sampling_Time[:19], "%Y-%m-%d %H:%M:%S")
    #print("date_object =", date_object)
    #print("type of date_object =", type(date_object))
    if (date_object.__class__.__name__ != 'datetime'):
        continue
    strSampling_Time = date_object.strftime("%Y-%m-%dT%H:%M:%S") + TZ
    Location = sheet.cell(row = i, column = 14).value #Location[mm]
    Depth = sheet.cell(row = i, column = 15).value #Depth[%]
    Lon_Trail = sheet.cell(row = i, column = 16).value #Lon_Trail[mm]
    short_diameter = sheet.cell(row = i, column = 17).value #short_diameter[mm]
    Bubble_Thickness = sheet.cell(row = i, column = 18).value #Bubble_Thickness[mm]
    Volume = sheet.cell(row = i, column = 19).value #Volume[nL]
    Lon_Trail_Log = sheet.cell(row = i, column = 20).value #Lon_Trail_Log
    Internal_Pressure_Log = sheet.cell(row = i, column = 21).value #Internal_Pressure_Log
    Internal_Pressure = sheet.cell(row = i, column = 22).value #Internal_Pressure[kPa]
    N2 = sheet.cell(row = i, column = 23).value #N2[%]
    CO2 = sheet.cell(row = i, column = 24).value #CO2[%]
    O2 = sheet.cell(row = i, column = 25).value #O2[%]
    Ar = sheet.cell(row = i, column = 26).value #Ar[%]
    SO2 = sheet.cell(row = i, column = 27).value #SO2[%]
    cnt=cnt+1
    print(i, ': AnaStatus:', AnaStatus ,': Sampling_Time:', Sampling_Time,': N2:', N2,': Ar:', Ar)
    sql = "INSERT INTO BubbleAnalysis(Sampling_Time,Location,Depth,Lon_Trail,short_diameter,Bubble_Thickness,Volume,Lon_Trail_Log,Internal_Pressure_Log,Internal_Pressure,N2,CO2,O2,Ar,SO2) VALUES ( "
    sql = sql + " '" +  strSampling_Time  + "','" + str(Location) + "','" + str(Depth) + "',"
    sql = sql + " '" + str(Lon_Trail) + "','" + str(short_diameter) + "','" + str(Internal_Pressure_Log) + "',"
    sql = sql + " '" + str(Volume) + "','" + str(Lon_Trail_Log) + "','" + str(Bubble_Thickness) + "',"
    sql = sql + " '" + str(Internal_Pressure) + "','" + str(N2) + "','" + str(CO2) + "',"
    sql = sql + " '" + str(O2) + "','" + str(Ar) + "','" + str(SO2) + "')"
    #print(sql)
    c1.execute(sql)
if(cnt==0):
    log2file("No Fresh Record!!")
    sendmail('No Fresh Record!!')
    exit(0)
Config.set("LastFileInfo","TestDate", TestDatestr) 
# 寫入 INI 檔案
with open('pyxl2ai.ini', 'w',encoding='UTF8') as configfile:
    Config.write(configfile)
log2file("Data insert into to sqlite complete!!")

# stastic data export for DataBaseServer
csvPath = FileInfo['outputdir']
outputfilename = FileInfo['outputfilename'] + now.strftime("%Y%m%d") + ".csv"
csvfilname = csvPath + outputfilename

headlist=['Timestamp','Location','Depth','Lon_Trail','short_diameter','Bubble_Thickness','Volume','Lon_Trail_Log','Internal_Pressure_Log','Internal_Pressure','N2','CO2','O2','Ar','SO2']
sql = "SELECT * FROM BubbleAnalysis"
y1 = c1.execute(sql)
with open(csvfilname, 'w', newline='') as csvfile:
    writer = csv.writer(csvfile, quoting=csv.QUOTE_ALL,delimiter=',')
    writer.writerow(headlist)
    writer.writerows(y1)
print("Generate CSV file complete")

log2file("execute complete")
os.remove(Tempfile)
sftp(outputfilename)

# if exeStatus:
sendmail('Process Successful')
print('Process complete')
# else :
#     sendmail('Process Fail')